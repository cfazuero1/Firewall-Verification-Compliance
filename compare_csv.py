import warnings
warnings.filterwarnings("ignore", message=".*does not match any known type.*")

import pandas as pd
from difflib import SequenceMatcher
from openpyxl import Workbook
from openpyxl.styles import Font
import argparse

# Function to calculate similarity between two strings
def similar(a, b):
    return SequenceMatcher(None, a, b).ratio()

# Function to apply color to cells
def apply_color(ws, cell, font_color, text):
    ws[cell].value = text
    ws[cell].font = Font(color=font_color)

# Function to compare two CSV files and generate an Excel file with colored differences
def compare_csv_excel(file1, file2, output_file, similarity_threshold=0.7):
    # Load both CSV files into DataFrames
    df1 = pd.read_csv(file1)
    df2 = pd.read_csv(file2)

    # Ensure both CSVs have the same columns
    if df1.columns.tolist() != df2.columns.tolist():
        raise ValueError("The two CSV files have different columns")

    # Create a new Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison Results"

    # Write the header row
    header = ['File1 Row', 'File2 Row', 'Similarity'] + df1.columns.tolist() + [col + " (File2)" for col in df1.columns]
    ws.append(header)

    # Iterate over rows of both DataFrames and compare each row
    for index1, row1 in df1.iterrows():
        for index2, row2 in df2.iterrows():
            # Compare rows based on similarity of all columns
            row_similarity = [similar(str(row1[col]), str(row2[col])) for col in df1.columns]
            avg_similarity = sum(row_similarity) / len(row_similarity)

            # If rows are sufficiently similar, check for differences
            if avg_similarity >= similarity_threshold:
                differences = []
                for col in df1.columns:
                    differences.append((row1[col], row2[col]))

                # Append the row information to the worksheet
                ws.append([index1, index2, avg_similarity])

                # Apply color highlighting for differences and leave matches uncolored
                for col_index, diff in enumerate(differences, start=4):
                    cell1_address = ws.cell(row=ws.max_row, column=col_index).coordinate
                    cell2_address = ws.cell(row=ws.max_row, column=col_index + len(df1.columns)).coordinate

                    if diff[0] != diff[1]:  # Only apply color to differences
                        apply_color(ws, cell1_address, '8B0000', str(diff[0]))  # Dark Red for file1 value
                        apply_color(ws, cell2_address, '006400', str(diff[1]))  # Dark Green for file2 value
                    else:
                        ws[cell1_address].value = diff[0]  # No color for matching values
                        ws[cell2_address].value = diff[1]

    # Save the Excel file
    wb.save(output_file)
    print(f"Comparison results with colored differences saved to {output_file}")

# Main function to handle command-line arguments
def main():
    parser = argparse.ArgumentParser(description="Compare two CSV files and highlight differences in Excel.")
    parser.add_argument('file1', help="Path to the first CSV file.")
    parser.add_argument('file2', help="Path to the second CSV file.")
    parser.add_argument('-o', '--output', default='comparison_results.xlsx', help="Path to the output Excel file.")
    parser.add_argument('-t', '--threshold', type=float, default=0.7, help="Similarity threshold (default: 0.7).")

    args = parser.parse_args()

    compare_csv_excel(args.file1, args.file2, args.output, args.threshold)

if __name__ == "__main__":
    main()
